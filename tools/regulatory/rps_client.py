"""Berkeley Lab RPS workbook parser."""

from __future__ import annotations

import logging
from pathlib import Path
from typing import Dict, Iterable, Optional

from openpyxl import load_workbook

import config

logger = logging.getLogger(__name__)

_TARGETS_FILE = "RPS-CES-Targets-and-Demand-*.xlsx"
_ADDITIONS_FILE = "RPS-CES-Capacity-Additions-*.xlsx"
_COSTS_FILE = "RPS-CES-Compliance-Costs-*.xlsx"


def _default_workbook_dir() -> Path:
    reg_cfg = getattr(config, "REGULATORY", {})
    return Path(reg_cfg.get("rps_workbook_dir", "data"))


def _find_one(directory: Path, pattern: str) -> Optional[Path]:
    candidates = sorted(directory.glob(pattern))
    return candidates[-1] if candidates else None


def _find_header_row(sheet) -> Optional[int]:
    for row_idx in range(1, min(sheet.max_row, 40) + 1):
        values = [sheet.cell(row=row_idx, column=col).value for col in range(1, min(sheet.max_column, 12) + 1)]
        text_values = {str(v).strip() for v in values if v not in (None, "")}
        if "State" in text_values and any(isinstance(v, int) and 2000 <= v <= 2050 for v in values):
            return row_idx
    return None


def _iter_wide_rows(sheet) -> Iterable[dict]:
    header_row = _find_header_row(sheet)
    if header_row is None:
        return []

    headers = [sheet.cell(row=header_row, column=col).value for col in range(1, sheet.max_column + 1)]
    rows = []
    for row_idx in range(header_row + 1, sheet.max_row + 1):
        row = {headers[col - 1]: sheet.cell(row=row_idx, column=col).value for col in range(1, sheet.max_column + 1)}
        state = str(row.get("State") or "").strip()
        if not state:
            continue
        rows.append(row)
    return rows


def _standard_type_for_tier(tier: str) -> str:
    return "CES" if "CES" in tier.upper() else "RPS"


def _ensure_record(records: dict, state: str, tier: str, year: int, notes: str) -> dict:
    key = (state, tier or "Total RPS", year)
    record = records.get(key)
    if record is None:
        record = {
            "state": state,
            "standard_type": _standard_type_for_tier(tier or "Total RPS"),
            "tier": tier or "Total RPS",
            "year": year,
            "target_pct": None,
            "demand_gwh": None,
            "applicable_sales_gwh": None,
            "statewide_sales_gwh": None,
            "achievement_ratio": None,
            "compliance_cost_per_kwh": None,
            "capacity_additions_mw": None,
            "notes": notes or "",
            "properties": {},
        }
        records[key] = record
    if notes and not record.get("notes"):
        record["notes"] = notes
    return record


def _coerce_float(value) -> Optional[float]:
    if value in (None, ""):
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def _load_targets_and_demand(path: Path, records: dict) -> None:
    wb = load_workbook(path, data_only=True, read_only=True)
    sheet_map = {
        "Statewide Sales": "statewide_sales_gwh",
        "RPS-Applicable Sales": "applicable_sales_gwh",
        "RPS & CES Targets (%)": "target_pct",
        "RPS & CES Demand (GWh)": "demand_gwh",
    }
    for sheet_name, field in sheet_map.items():
        if sheet_name not in wb.sheetnames:
            continue
        for row in _iter_wide_rows(wb[sheet_name]):
            state = str(row.get("State") or "").strip()
            notes = str(row.get("Special Notes") or "").strip()
            tier = str(row.get("RPS Tier or Carve Out") or "Total RPS").strip()
            for year, value in row.items():
                if not isinstance(year, int):
                    continue
                parsed = _coerce_float(value)
                if parsed is None:
                    continue
                record = _ensure_record(records, state, tier, year, notes)
                record[field] = parsed


def _load_capacity_additions(path: Path, records: dict) -> None:
    wb = load_workbook(path, data_only=True, read_only=True)
    if "RPS & CES Capacity Additions" not in wb.sheetnames:
        return
    for row in _iter_wide_rows(wb["RPS & CES Capacity Additions"]):
        state = str(row.get("State") or "").strip()
        technology = str(row.get("Technology") or "").strip()
        for year, value in row.items():
            if not isinstance(year, int):
                continue
            parsed = _coerce_float(value)
            if parsed is None:
                continue
            record = _ensure_record(records, state, "Total RPS", year, "")
            current = record.get("capacity_additions_mw") or 0.0
            record["capacity_additions_mw"] = current + parsed
            by_tech = record["properties"].setdefault("capacity_additions_by_technology", {})
            by_tech[technology] = float(by_tech.get(technology, 0.0)) + parsed


def _load_costs_and_achievement(path: Path, records: dict) -> None:
    wb = load_workbook(path, data_only=True, read_only=True)
    sheet_map = {
        "Target Achievement": "achievement_ratio",
        "Compliance Costs": "compliance_cost_per_kwh",
    }
    for sheet_name, field in sheet_map.items():
        if sheet_name not in wb.sheetnames:
            continue
        for row in _iter_wide_rows(wb[sheet_name]):
            state = str(row.get("State") or "").strip()
            notes = str(row.get("Notes") or "").strip()
            tier = str(row.get("Total RPS or Tier") or "Total RPS").strip()
            for year, value in row.items():
                if not isinstance(year, int):
                    continue
                parsed = _coerce_float(value)
                if parsed is None:
                    continue
                record = _ensure_record(records, state, tier, year, notes)
                record[field] = parsed
                sources = str(row.get("Data Source(s)") or "").strip()
                if sources:
                    record["properties"]["data_sources"] = sources


def load_rps_data(workbook_dir: Optional[str] = None) -> list[dict]:
    """Parse the Berkeley Lab RPS workbooks into flat per-state/year rows."""
    directory = Path(workbook_dir).expanduser() if workbook_dir else _default_workbook_dir()
    targets_path = _find_one(directory, _TARGETS_FILE)
    additions_path = _find_one(directory, _ADDITIONS_FILE)
    costs_path = _find_one(directory, _COSTS_FILE)
    if not targets_path or not additions_path or not costs_path:
        logger.warning("RPS workbooks missing in %s", directory)
        return []

    records: Dict[tuple, dict] = {}
    try:
        _load_targets_and_demand(targets_path, records)
        _load_capacity_additions(additions_path, records)
        _load_costs_and_achievement(costs_path, records)
    except Exception as exc:
        logger.warning("Failed parsing RPS workbooks: %s", exc)
        return []

    return [
        records[key]
        for key in sorted(records.keys(), key=lambda item: (item[0], item[2], item[1]))
    ]


def latest_workbook_mtime(workbook_dir: Optional[str] = None) -> Optional[float]:
    """Return the latest mtime across the three expected workbooks."""
    directory = Path(workbook_dir).expanduser() if workbook_dir else _default_workbook_dir()
    paths = [
        _find_one(directory, _TARGETS_FILE),
        _find_one(directory, _ADDITIONS_FILE),
        _find_one(directory, _COSTS_FILE),
    ]
    existing = [path.stat().st_mtime for path in paths if path]
    return max(existing) if existing else None
