"""Eskom tariff schedule parser -- reads the Schedule of Standard Prices xlsm."""

from __future__ import annotations

import logging
from pathlib import Path
from typing import Any, Dict, List, Optional

logger = logging.getLogger(__name__)

_cache: Dict[str, Any] = {}

# Sheets to skip (not tariff rate tables)
_SKIP_SHEETS = {
    "Menu", "Loss Factors", "Gen Reconciliatons", "Gen-offset",
    " TUoS NLA", "DUoS NLA", " Excess NCC NLA", "Excess NCC Munic",
}

# Column indices (0-based) for rate data in tariff sheets
_COL_TARIFF = 0
_COL_TX_ZONE = 2
_COL_VOLTAGE = 3
_COL_ZONE_LABEL = 7
_COL_VOLTAGE_LABEL = 8
_COL_HD_PEAK = 9
_COL_HD_PEAK_VAT = 10
_COL_HD_STANDARD = 11
_COL_HD_STANDARD_VAT = 12
_COL_HD_OFFPEAK = 13
_COL_HD_OFFPEAK_VAT = 14
_COL_LD_PEAK = 15
_COL_LD_PEAK_VAT = 16
_COL_LD_STANDARD = 17
_COL_LD_STANDARD_VAT = 18
_COL_LD_OFFPEAK = 19
_COL_LD_OFFPEAK_VAT = 20
_COL_LEGACY = 21
_COL_LEGACY_VAT = 22
_COL_GEN_CAPACITY = 23
_COL_GEN_CAPACITY_VAT = 24
_COL_TX_NETWORK = 25
_COL_TX_NETWORK_VAT = 26


def _safe_float(val: Any) -> Optional[float]:
    if val is None:
        return None
    try:
        return float(val)
    except (ValueError, TypeError):
        return None


def _safe_int(val: Any) -> Optional[int]:
    if val is None:
        return None
    try:
        return int(val)
    except (ValueError, TypeError):
        return None


def _parse_workbook(file_path: str) -> Dict[str, List[dict]]:
    """Parse all tariff sheets from the Eskom tariff workbook."""
    try:
        import openpyxl
    except ImportError:
        logger.error("openpyxl is required to parse Eskom tariff files")
        return {}

    try:
        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    except Exception as exc:
        logger.error("Failed to open Eskom tariff file %s: %s", file_path, exc)
        return {}

    tariffs: Dict[str, List[dict]] = {}
    last_zone_label = ""

    for sheet_name in wb.sheetnames:
        if sheet_name.strip() in _SKIP_SHEETS:
            continue

        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))

        for row_idx, row in enumerate(rows):
            if row_idx < 7:  # Data starts at row 8 (0-indexed row 7)
                continue
            if len(row) < 21:
                continue

            tariff_name = row[_COL_TARIFF]
            if tariff_name is None or not str(tariff_name).strip():
                continue

            tariff_name = str(tariff_name).strip()
            tx_zone = _safe_int(row[_COL_TX_ZONE])
            voltage = _safe_int(row[_COL_VOLTAGE])

            if tx_zone is None or voltage is None:
                continue

            zone_label = row[_COL_ZONE_LABEL]
            if zone_label is not None and str(zone_label).strip():
                last_zone_label = str(zone_label).strip().replace("\n", " ")

            voltage_label = row[_COL_VOLTAGE_LABEL]
            voltage_label = str(voltage_label).strip() if voltage_label else ""

            rate_row = {
                "tariff": tariff_name,
                "sheet": sheet_name.strip(),
                "tx_zone": tx_zone,
                "tx_zone_label": last_zone_label,
                "voltage": voltage,
                "voltage_label": voltage_label,
                "high_demand_peak": _safe_float(row[_COL_HD_PEAK]),
                "high_demand_peak_vat": _safe_float(row[_COL_HD_PEAK_VAT]),
                "high_demand_standard": _safe_float(row[_COL_HD_STANDARD]),
                "high_demand_standard_vat": _safe_float(row[_COL_HD_STANDARD_VAT]),
                "high_demand_offpeak": _safe_float(row[_COL_HD_OFFPEAK]),
                "high_demand_offpeak_vat": _safe_float(row[_COL_HD_OFFPEAK_VAT]),
                "low_demand_peak": _safe_float(row[_COL_LD_PEAK]),
                "low_demand_peak_vat": _safe_float(row[_COL_LD_PEAK_VAT]),
                "low_demand_standard": _safe_float(row[_COL_LD_STANDARD]),
                "low_demand_standard_vat": _safe_float(row[_COL_LD_STANDARD_VAT]),
                "low_demand_offpeak": _safe_float(row[_COL_LD_OFFPEAK]),
                "low_demand_offpeak_vat": _safe_float(row[_COL_LD_OFFPEAK_VAT]),
                "legacy_charge": _safe_float(row[_COL_LEGACY]),
                "legacy_charge_vat": _safe_float(row[_COL_LEGACY_VAT]),
                "gen_capacity_charge": _safe_float(row[_COL_GEN_CAPACITY]),
                "gen_capacity_charge_vat": _safe_float(row[_COL_GEN_CAPACITY_VAT]),
                "tx_network_charge": _safe_float(row[_COL_TX_NETWORK]),
                "tx_network_charge_vat": _safe_float(row[_COL_TX_NETWORK_VAT]),
            }

            tariffs.setdefault(tariff_name, []).append(rate_row)

    wb.close()
    return tariffs


def _get_parsed(file_path: Optional[str] = None) -> Dict[str, List[dict]]:
    """Return parsed tariff data, using cache if available."""
    if file_path is None:
        try:
            import config
            cfg = getattr(config, "ESKOM", {})
            file_path = str(
                Path(cfg.get("data_dir", "data")) / cfg.get(
                    "tariff_file",
                    "Eskom-tariffs-1-April-2025-ver-2.xlsm",
                )
            )
        except ImportError:
            file_path = "data/Eskom-tariffs-1-April-2025-ver-2.xlsm"

    if file_path not in _cache:
        _cache[file_path] = _parse_workbook(file_path)

    return _cache[file_path]


def list_tariff_names(file_path: Optional[str] = None) -> List[str]:
    """Return list of available tariff names."""
    return sorted(_get_parsed(file_path).keys())


def get_tariff_rates(
    tariff_name: str,
    tx_zone: Optional[int] = None,
    voltage: Optional[int] = None,
    file_path: Optional[str] = None,
) -> List[dict]:
    """Get tariff rates, optionally filtered by zone and voltage."""
    data = _get_parsed(file_path)
    rows = data.get(tariff_name, [])

    if tx_zone is not None:
        rows = [r for r in rows if r["tx_zone"] == tx_zone]
    if voltage is not None:
        rows = [r for r in rows if r["voltage"] == voltage]

    return rows


def get_rate(
    tariff_name: str,
    tx_zone: int,
    voltage: int,
    season: str,
    period: str,
    file_path: Optional[str] = None,
) -> Optional[float]:
    """Get a single rate value in c/kWh for the specified parameters.

    Args:
        tariff_name: e.g. ``"Megaflex"``
        tx_zone: 0-3
        voltage: 1-4
        season: ``"high"`` or ``"low"``
        period: ``"peak"``, ``"standard"``, or ``"offpeak"``
        file_path: path to tariff workbook (optional)

    Returns:
        Rate in c/kWh (excl VAT), or None if not found.
    """
    rows = get_tariff_rates(tariff_name, tx_zone=tx_zone, voltage=voltage,
                            file_path=file_path)
    if not rows:
        return None

    key = f"{season.lower()}_demand_{period.lower()}"
    return rows[0].get(key)
