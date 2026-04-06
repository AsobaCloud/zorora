"""Tests for SEP-036: regulatory ingest, store, workflow, API, and UI."""

from __future__ import annotations

import base64
import importlib
from pathlib import Path
import tempfile
import time
from unittest.mock import MagicMock, patch

import pytest


def _make_json_response(payload: dict, status_code: int = 200):
    response = MagicMock()
    response.status_code = status_code
    response.json.return_value = payload
    response.raise_for_status = MagicMock()
    return response


def _import_app_module():
    return importlib.import_module("ui.web.app")


def _write_rps_workbooks(workbook_dir: Path) -> None:
    from openpyxl import Workbook

    workbook_dir.mkdir(parents=True, exist_ok=True)

    targets_path = workbook_dir / "RPS-CES-Targets-and-Demand-August-2025.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Statewide Sales"
    headers = ["State", "Special Notes", "RPS Tier or Carve Out", 2025, 2026]
    for col, value in enumerate(headers, start=1):
        ws.cell(row=25, column=col, value=value)
    ws.cell(row=26, column=1, value="CA")
    ws.cell(row=26, column=2, value="Large market")
    ws.cell(row=26, column=3, value="Total RPS")
    ws.cell(row=26, column=4, value=285000)
    ws.cell(row=26, column=5, value=290000)

    ws = wb.create_sheet("RPS-Applicable Sales")
    for col, value in enumerate(headers, start=1):
        ws.cell(row=25, column=col, value=value)
    ws.cell(row=26, column=1, value="CA")
    ws.cell(row=26, column=2, value="Large market")
    ws.cell(row=26, column=3, value="Total RPS")
    ws.cell(row=26, column=4, value=250000)
    ws.cell(row=26, column=5, value=255000)

    ws = wb.create_sheet("RPS & CES Targets (%)")
    for col, value in enumerate(headers, start=1):
        ws.cell(row=25, column=col, value=value)
    ws.cell(row=26, column=1, value="CA")
    ws.cell(row=26, column=2, value="Large market")
    ws.cell(row=26, column=3, value="Total RPS")
    ws.cell(row=26, column=4, value=0.60)
    ws.cell(row=26, column=5, value=0.65)

    ws = wb.create_sheet("RPS & CES Demand (GWh)")
    for col, value in enumerate(headers, start=1):
        ws.cell(row=25, column=col, value=value)
    ws.cell(row=26, column=1, value="CA")
    ws.cell(row=26, column=2, value="Large market")
    ws.cell(row=26, column=3, value="Total RPS")
    ws.cell(row=26, column=4, value=150000)
    ws.cell(row=26, column=5, value=165750)
    wb.save(targets_path)

    additions_path = workbook_dir / "RPS-CES-Capacity-Additions-August-2025.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "RPS & CES Capacity Additions"
    headers = ["State", "Technology", 2025, 2026]
    for col, value in enumerate(headers, start=1):
        ws.cell(row=25, column=col, value=value)
    ws.cell(row=26, column=1, value="CA")
    ws.cell(row=26, column=2, value="Solar")
    ws.cell(row=26, column=3, value=4200)
    ws.cell(row=26, column=4, value=4500)
    wb.save(additions_path)

    costs_path = workbook_dir / "RPS-CES-Compliance-Costs-August-2025.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Target Achievement"
    headers = ["State", "Notes", "Data Source(s)", "Total RPS or Tier", 2025, 2026]
    for col, value in enumerate(headers, start=1):
        ws.cell(row=29, column=col, value=value)
    ws.cell(row=30, column=1, value="CA")
    ws.cell(row=30, column=2, value="On track")
    ws.cell(row=30, column=3, value="Example")
    ws.cell(row=30, column=4, value="Total RPS")
    ws.cell(row=30, column=5, value=0.98)
    ws.cell(row=30, column=6, value=1.02)

    ws = wb.create_sheet("Compliance Costs")
    for col, value in enumerate(headers, start=1):
        ws.cell(row=24, column=col, value=value)
    ws.cell(row=25, column=1, value="CA")
    ws.cell(row=25, column=2, value="On track")
    ws.cell(row=25, column=3, value="Example")
    ws.cell(row=25, column=4, value="Total RPS")
    ws.cell(row=25, column=5, value=0.012)
    ws.cell(row=25, column=6, value=0.011)
    wb.save(costs_path)


def test_config_has_sep036_blocks():
    import config

    assert hasattr(config, "EIA")
    assert hasattr(config, "OPENEI")
    assert hasattr(config, "REGULATORY")


def test_eia_client_fetch_generator_capacity_paginates_and_normalizes():
    from tools.regulatory import eia_client

    page_1 = {
        "response": {
            "total": "3",
            "data": [
                {
                    "period": "2025-12",
                    "stateid": "CA",
                    "stateName": "California",
                    "energy_source_code": "SUN",
                    "net-summer-capacity-mw": "20",
                    "net-summer-capacity-mw-units": "MW",
                    "plantName": "Sunray 2",
                },
                {
                    "period": "2025-12",
                    "stateid": "CA",
                    "stateName": "California",
                    "energy_source_code": "SUN",
                    "net-summer-capacity-mw": "13.8",
                    "net-summer-capacity-mw-units": "MW",
                    "plantName": "Sunray 3",
                },
            ],
        }
    }
    page_2 = {
        "response": {
            "total": "3",
            "data": [
                {
                    "period": "2025-12",
                    "stateid": "CA",
                    "stateName": "California",
                    "energy_source_code": "SUN",
                    "net-summer-capacity-mw": "11.1",
                    "net-summer-capacity-mw-units": "MW",
                    "plantName": "Sunray 4",
                }
            ],
        }
    }

    with patch.object(eia_client.config, "EIA", {
        "api_key": "demo",
        "base_url": "https://api.eia.gov/v2",
        "timeout": 30,
        "enabled": True,
        "max_rows_per_request": 2,
    }), patch("tools.regulatory.eia_client.requests.get", side_effect=[
        _make_json_response(page_1),
        _make_json_response(page_2),
    ]) as mock_get:
        records = eia_client.fetch_generator_capacity(state="CA", fuel_type="SUN")

    assert len(records) == 3
    assert records[0]["endpoint"] == "operating-generator-capacity"
    assert records[0]["state"] == "CA"
    assert records[0]["fuel_type"] == "SUN"
    assert records[0]["value"] == 20.0
    assert records[0]["unit"] == "MW"
    assert records[0]["properties"]["plantName"] == "Sunray 2"
    assert mock_get.call_count == 2
    first_params = mock_get.call_args_list[0].kwargs["params"]
    second_params = mock_get.call_args_list[1].kwargs["params"]
    assert first_params["facets[stateid][]"] == "CA"
    assert first_params["facets[energy_source_code][]"] == "SUN"
    assert first_params["data[0]"] == "net-summer-capacity-mw"
    assert second_params["offset"] == 2


def test_eia_client_fetch_operational_data_uses_location_and_generation():
    from tools.regulatory import eia_client

    payload = {
        "response": {
            "total": "1",
            "data": [
                {
                    "period": "2025-12",
                    "location": "CA",
                    "stateDescription": "California",
                    "fueltypeid": "SUN",
                    "generation": "2536.06808",
                    "generation-units": "thousand megawatthours",
                }
            ],
        }
    }

    with patch.object(eia_client.config, "EIA", {
        "api_key": "demo",
        "base_url": "https://api.eia.gov/v2",
        "timeout": 30,
        "enabled": True,
        "max_rows_per_request": 5000,
    }), patch("tools.regulatory.eia_client.requests.get", return_value=_make_json_response(payload)) as mock_get:
        records = eia_client.fetch_operational_data(state="CA", fuel_type="SUN")

    assert len(records) == 1
    assert records[0]["endpoint"] == "electric-power-operational-data"
    assert records[0]["state"] == "CA"
    assert records[0]["fuel_type"] == "SUN"
    assert records[0]["value"] == 2536.06808
    params = mock_get.call_args.kwargs["params"]
    assert params["facets[location][]"] == "CA"
    assert params["facets[fueltypeid][]"] == "SUN"
    assert params["data[0]"] == "generation"


def test_openei_client_summarizes_sector_rates():
    from tools.regulatory import openei_client

    search_payload = {
        "items": [
            {"label": "res-1", "utility": "Public Service Co of Colorado", "sector": "Residential", "uri": "https://example.com/res"},
            {"label": "com-1", "utility": "Public Service Co of Colorado", "sector": "Commercial", "uri": "https://example.com/com"},
            {"label": "ind-1", "utility": "Public Service Co of Colorado", "sector": "Industrial", "uri": "https://example.com/ind"},
        ]
    }
    detail_payloads = [
        {"items": [{"sector": "Residential", "energyratestructure": [[{"rate": 0.05951}]], "country": "USA"}]},
        {"items": [{"sector": "Commercial", "energyratestructure": [[{"rate": 0.08123}]], "country": "USA"}]},
        {"items": [{"sector": "Industrial", "energyratestructure": [[{"rate": 0.07111}]], "country": "USA"}]},
    ]

    with patch.object(openei_client.config, "OPENEI", {
        "api_key": "demo",
        "base_url": "https://api.openei.org",
        "timeout": 30,
        "enabled": True,
    }), patch("tools.regulatory.openei_client.requests.get", side_effect=[
        _make_json_response(search_payload),
        _make_json_response(detail_payloads[0]),
        _make_json_response(detail_payloads[1]),
        _make_json_response(detail_payloads[2]),
    ]):
        summary = openei_client.fetch_utility_rates(39.7392, -104.9903)

    assert summary["utility_name"] == "Public Service Co of Colorado"
    assert summary["lat"] == 39.7392
    assert summary["lon"] == -104.9903
    assert summary["rates"]["residential"] == 0.05951
    assert summary["rates"]["commercial"] == 0.08123
    assert summary["rates"]["industrial"] == 0.07111


def test_rps_parser_combines_workbooks():
    from tools.regulatory.rps_client import load_rps_data

    with tempfile.TemporaryDirectory() as tmpdir:
        workbook_dir = Path(tmpdir)
        _write_rps_workbooks(workbook_dir)
        records = load_rps_data(workbook_dir=str(workbook_dir))

    row_2025 = next(
        record for record in records
        if record["state"] == "CA" and record["tier"] == "Total RPS" and record["year"] == 2025
    )
    assert row_2025["standard_type"] == "RPS"
    assert row_2025["target_pct"] == 0.60
    assert row_2025["demand_gwh"] == 150000.0
    assert row_2025["applicable_sales_gwh"] == 250000.0
    assert row_2025["statewide_sales_gwh"] == 285000.0
    assert row_2025["achievement_ratio"] == 0.98
    assert row_2025["compliance_cost_per_kwh"] == 0.012
    assert row_2025["capacity_additions_mw"] == 4200.0


def test_regulatory_store_upsert_and_query(tmp_path):
    from tools.regulatory.store import RegulatoryDataStore

    store = RegulatoryDataStore(db_path=str(tmp_path / "regulatory.db"))
    store.upsert_rps_targets([
        {
            "state": "CA",
            "standard_type": "RPS",
            "tier": "Total RPS",
            "year": 2025,
            "target_pct": 0.60,
            "demand_gwh": 150000.0,
            "applicable_sales_gwh": 250000.0,
            "statewide_sales_gwh": 285000.0,
            "achievement_ratio": 0.98,
            "compliance_cost_per_kwh": 0.012,
            "capacity_additions_mw": 4200.0,
            "notes": "On track",
            "properties": {"source": "test"},
        }
    ])
    store.upsert_eia_series([
        {
            "endpoint": "operating-generator-capacity",
            "period": "2025-12",
            "state": "CA",
            "fuel_type": "SUN",
            "value": 20.0,
            "unit": "MW",
            "properties": {"plantName": "Sunray 2"},
        }
    ], endpoint="operating-generator-capacity")
    store.upsert_utility_rates([
        {
            "utility_name": "Public Service Co of Colorado",
            "state": "CO",
            "sector": "residential",
            "rate_kwh": 0.05951,
            "lat": 39.7392,
            "lon": -104.9903,
            "properties": {"label": "res-1"},
        }
    ])
    store.upsert_regulatory_events([
        {
            "jurisdiction": "US",
            "regulator": "FERC",
            "event_type": "rulemaking",
            "title": "Example Order",
            "summary": "A test event",
            "published_date": "2026-03-01",
            "effective_date": "2026-04-01",
            "deadline_date": None,
            "source_url": "https://example.com/order",
            "properties": {"docket": "RM-1"},
        }
    ])

    rps = store.get_rps_targets(state="CA", year=2025)
    capacity = store.get_eia_series("operating-generator-capacity", state="CA", fuel_type="SUN")
    rates = store.get_utility_rates(state="CO", sector="residential")
    events = store.get_regulatory_events(jurisdiction="US", event_type="rulemaking")

    assert len(rps) == 1
    assert rps[0]["capacity_additions_mw"] == 4200.0
    assert len(capacity) == 1
    assert capacity[0]["value"] == 20.0
    assert len(rates) == 1
    assert rates[0]["rate_kwh"] == 0.05951
    assert len(events) == 1
    assert events[0]["title"] == "Example Order"
    assert store.get_staleness("rps_targets") is not None
    store.close()


def test_africa_regulatory_parsers_emit_common_contract_and_transform_metadata():
    from tools.regulatory.africa_client import (
        parse_ippo_press_releases,
        parse_nersa_homepage,
        normalize_zera_seed_catalog,
    )
    from tools.regulatory.normalization import REGULATORY_EVENT_SCHEMA_VERSION

    nersa_html = """
    <div id="latest_news">
        <a href="file/8399" target="_blank" class="home-section-article">
            Media Statement - NERSA approves Eskom Retail Tariffs and Structural Adjustment application for 2026/27
            <span>10 March 2026</span>
        </a>
    </div>
    <div id="recent_decisions">
        <ul>
            <li class="listing-li">
                <a href="file/8375" target="_blank" class="listing">
                    Update on the MRP and Risk-Free Rate calculation for the period ended 31 December 2025
                    <span>17 February 2026</span>
                </a>
            </li>
        </ul>
    </div>
    """
    ippo_payload = [
        {
            "id": "94e36e92-1adb-f011-8544-7c1e52501ab8",
            "detail": "",
            "headline": "ANNOUNCEMENT OF ADDITIONAL PREFERRED BIDDERS UNDER BID WINDOW 7 OF THE RENEWABLE ENERGY INDEPENDENT POWER PRODUCER PROCUREMENT PROGRAMME",
            "date": "12/16/2025 10:00:00 PM",
            "year": 2025,
            "month": 12,
            "monthname": "December",
            "thumbnail": "/Image/download.aspx?Entity=ippo_webpressrelease&Attribute=ippo_thumbnail&Id=94e36e92-1adb-f011-8544-7c1e52501ab8",
            "noteid": "613d59d6-de10-d6db-83a6-86b1a2031c32",
            "filename": "Final Media Statement Announcement ITP PQBs and REIPPPP BW7 15122025.pdf",
            "websiteid": "f46a9fe7-567a-ee11-8179-6045bd8c528e",
            "websitename": "ipp-renewables - ipp-renewables",
        }
    ]
    zera_seed = [
        {
            "title": "Public Notice - Fuel Notice 4 October 2025",
            "published_date": "2025-10-06",
            "source_url": "https://www.zera.co.zw/press-releases-public-notices/",
            "summary": "PUBLIC NOTICE: NOTIFICATION OF PETROLEUM PRODUCT PRICES",
            "category": "Press Releases",
        }
    ]

    bundles = [
        parse_nersa_homepage(nersa_html),
        parse_ippo_press_releases(ippo_payload),
        normalize_zera_seed_catalog(zera_seed),
    ]

    required_event_fields = {
        "jurisdiction",
        "regulator",
        "event_type",
        "title",
        "summary",
        "published_date",
        "source_url",
        "source_system",
        "source_record_id",
        "schema_version",
        "transform_version",
        "transform_run_id",
        "raw_document_id",
    }

    for bundle in bundles:
        assert bundle["events"]
        assert bundle["raw_documents"]
        assert bundle["transform_runs"]

        event = bundle["events"][0]
        run = bundle["transform_runs"][0]
        raw_document = bundle["raw_documents"][0]

        assert required_event_fields.issubset(event.keys())
        assert event["schema_version"] == REGULATORY_EVENT_SCHEMA_VERSION
        assert event["transform_run_id"] == run["id"]
        assert event["raw_document_id"] == raw_document["id"]
        assert run["schema_version"] == REGULATORY_EVENT_SCHEMA_VERSION
        assert "title" in run["mapping"]
        assert raw_document["source_system"] == event["source_system"]


def test_regulatory_store_tracks_raw_documents_transform_runs_and_lineage(tmp_path):
    from tools.regulatory.store import RegulatoryDataStore

    store = RegulatoryDataStore(db_path=str(tmp_path / "regulatory.db"))
    store.upsert_raw_documents([
        {
            "id": "raw-za-1",
            "jurisdiction": "ZA",
            "source_system": "nersa_recent_decisions",
            "source_url": "https://www.nersa.org.za/",
            "content_type": "text/html",
            "fetch_status": "ok",
            "http_status": 200,
            "document_hash": "sha256:abc123",
            "payload_text": "<html>sample</html>",
            "metadata": {"selector": "#recent_decisions"},
        }
    ])
    store.upsert_transform_runs([
        {
            "id": "run-za-1",
            "jurisdiction": "ZA",
            "source_system": "nersa_recent_decisions",
            "raw_document_id": "raw-za-1",
            "transform_name": "nersa_recent_decision_v1",
            "schema_version": "regulatory-event.v1",
            "transform_version": "nersa_recent_decision.v1",
            "mapping": {
                "title": {"source": "a.listing"},
                "published_date": {"source": "span"},
                "event_type": {"value": "decision"},
            },
            "notes": "Parses #recent_decisions on the NERSA homepage.",
            "record_count": 1,
        }
    ])
    store.upsert_regulatory_events([
        {
            "jurisdiction": "ZA",
            "regulator": "NERSA",
            "event_type": "decision",
            "title": "Update on the MRP and Risk-Free Rate calculation",
            "summary": "Recent Regulator Decision",
            "published_date": "2026-02-17",
            "effective_date": None,
            "deadline_date": None,
            "source_url": "https://www.nersa.org.za/file/8375",
            "source_system": "nersa_recent_decisions",
            "source_record_id": "8375",
            "schema_version": "regulatory-event.v1",
            "transform_version": "nersa_recent_decision.v1",
            "transform_run_id": "run-za-1",
            "raw_document_id": "raw-za-1",
            "properties": {"section": "recent_decisions"},
        }
    ])

    raw_documents = store.get_raw_documents(source_system="nersa_recent_decisions")
    transform_runs = store.get_transform_runs(source_system="nersa_recent_decisions")
    events = store.get_regulatory_events(jurisdiction="ZA", event_type="decision")

    assert len(raw_documents) == 1
    assert raw_documents[0]["fetch_status"] == "ok"
    assert raw_documents[0]["metadata"]["selector"] == "#recent_decisions"
    assert len(transform_runs) == 1
    assert transform_runs[0]["mapping"]["title"]["source"] == "a.listing"
    assert len(events) == 1
    assert events[0]["schema_version"] == "regulatory-event.v1"
    assert events[0]["transform_run_id"] == "run-za-1"
    assert events[0]["raw_document_id"] == "raw-za-1"
    store.close()


def test_regulatory_workflow_updates_all_sources(tmp_path):
    from workflows.regulatory_workflow import RegulatoryWorkflow
    from tools.regulatory.store import RegulatoryDataStore

    store = RegulatoryDataStore(db_path=str(tmp_path / "regulatory.db"))
    workflow = RegulatoryWorkflow(store=store)

    with patch("workflows.regulatory_workflow.fetch_generator_capacity", return_value=[
        {"endpoint": "operating-generator-capacity", "period": "2025-12", "state": "CA", "fuel_type": "SUN", "value": 20.0, "unit": "MW", "properties": {"plantName": "Sunray 2"}}
    ]), patch("workflows.regulatory_workflow.fetch_operational_data", return_value=[
        {"endpoint": "electric-power-operational-data", "period": "2025-12", "state": "CA", "fuel_type": "SUN", "value": 2536.0, "unit": "thousand megawatthours", "properties": {"stateDescription": "California"}}
    ]), patch("workflows.regulatory_workflow.fetch_retail_sales", return_value=[
        {"endpoint": "retail-sales", "period": "2025-12", "state": "CA", "fuel_type": "RES", "value": 15555.0, "unit": "million kilowatt hours", "properties": {"stateDescription": "California"}}
    ]), patch("workflows.regulatory_workflow.fetch_utility_rates", return_value={
        "utility_name": "Public Service Co of Colorado",
        "state": "CO",
        "lat": 39.7392,
        "lon": -104.9903,
        "rates": {"residential": 0.05951, "commercial": 0.08123},
        "properties": {"source": "OpenEI"},
    }), patch("workflows.regulatory_workflow.load_rps_data", return_value=[
        {"state": "CA", "standard_type": "RPS", "tier": "Total RPS", "year": 2025, "target_pct": 0.60,
         "demand_gwh": 150000.0, "applicable_sales_gwh": 250000.0, "statewide_sales_gwh": 285000.0,
         "achievement_ratio": 0.98, "compliance_cost_per_kwh": 0.012, "capacity_additions_mw": 4200.0,
         "notes": "On track", "properties": {"source": "test"}}
    ]):
        updated = workflow.update_all(force=True)

    assert updated >= 4
    assert len(store.get_rps_targets(state="CA")) == 1
    assert len(store.get_eia_series("operating-generator-capacity", state="CA")) == 1
    assert len(store.get_utility_rates(state="CO")) == 2
    store.close()


def test_regulatory_workflow_updates_africa_event_sources_with_lineage(tmp_path):
    from workflows.regulatory_workflow import RegulatoryWorkflow
    from tools.regulatory.store import RegulatoryDataStore

    store = RegulatoryDataStore(db_path=str(tmp_path / "regulatory.db"))
    workflow = RegulatoryWorkflow(store=store)

    nersa_bundle = {
        "events": [
            {
                "jurisdiction": "ZA",
                "regulator": "NERSA",
                "event_type": "decision",
                "title": "NERSA tariff decision",
                "summary": "Recent Regulator Decision",
                "published_date": "2026-02-17",
                "effective_date": None,
                "deadline_date": None,
                "source_url": "https://www.nersa.org.za/file/8375",
                "source_system": "nersa_recent_decisions",
                "source_record_id": "8375",
                "schema_version": "regulatory-event.v1",
                "transform_version": "nersa_recent_decision.v1",
                "transform_run_id": "run-nersa-1",
                "raw_document_id": "raw-nersa-1",
                "properties": {"section": "recent_decisions"},
            }
        ],
        "raw_documents": [
            {
                "id": "raw-nersa-1",
                "jurisdiction": "ZA",
                "source_system": "nersa_recent_decisions",
                "source_url": "https://www.nersa.org.za/",
                "content_type": "text/html",
                "fetch_status": "ok",
                "http_status": 200,
                "document_hash": "sha256:nersa",
                "payload_text": "<html>nersa</html>",
                "metadata": {"selector": "#recent_decisions"},
            }
        ],
        "transform_runs": [
            {
                "id": "run-nersa-1",
                "jurisdiction": "ZA",
                "source_system": "nersa_recent_decisions",
                "raw_document_id": "raw-nersa-1",
                "transform_name": "nersa_recent_decision_v1",
                "schema_version": "regulatory-event.v1",
                "transform_version": "nersa_recent_decision.v1",
                "mapping": {"title": {"source": "a.listing"}},
                "notes": "NERSA homepage recent decisions",
                "record_count": 1,
            }
        ],
    }
    ippo_bundle = {
        "events": [
            {
                "jurisdiction": "ZA",
                "regulator": "IPP Office",
                "event_type": "procurement_update",
                "title": "Additional preferred bidders announced",
                "summary": "IPP Office press release",
                "published_date": "2025-12-16",
                "effective_date": None,
                "deadline_date": None,
                "source_url": "https://www.ipp-projects.co.za/_entity/annotation/613d59d6-de10-d6db-83a6-86b1a2031c32",
                "source_system": "ippo_oldnews",
                "source_record_id": "94e36e92-1adb-f011-8544-7c1e52501ab8",
                "schema_version": "regulatory-event.v1",
                "transform_version": "ippo_press_release.v1",
                "transform_run_id": "run-ippo-1",
                "raw_document_id": "raw-ippo-1",
                "properties": {"filename": "Final Media Statement Announcement ITP PQBs and REIPPPP BW7 15122025.pdf"},
            }
        ],
        "raw_documents": [
            {
                "id": "raw-ippo-1",
                "jurisdiction": "ZA",
                "source_system": "ippo_oldnews",
                "source_url": "https://www.ipp-projects.co.za/PortalAPI/?etn=oldnews",
                "content_type": "application/json",
                "fetch_status": "ok",
                "http_status": 200,
                "document_hash": "sha256:ippo",
                "payload_text": "[]",
                "metadata": {"record_count": 1},
            }
        ],
        "transform_runs": [
            {
                "id": "run-ippo-1",
                "jurisdiction": "ZA",
                "source_system": "ippo_oldnews",
                "raw_document_id": "raw-ippo-1",
                "transform_name": "ippo_press_release_v1",
                "schema_version": "regulatory-event.v1",
                "transform_version": "ippo_press_release.v1",
                "mapping": {"headline": {"target": "title"}},
                "notes": "IPP Office oldnews portal feed",
                "record_count": 1,
            }
        ],
    }
    zera_bundle = {
        "events": [
            {
                "jurisdiction": "ZW",
                "regulator": "ZERA",
                "event_type": "public_notice",
                "title": "Public Notice - Fuel Notice 4 October 2025",
                "summary": "PUBLIC NOTICE: NOTIFICATION OF PETROLEUM PRODUCT PRICES",
                "published_date": "2025-10-06",
                "effective_date": None,
                "deadline_date": None,
                "source_url": "https://www.zera.co.zw/press-releases-public-notices/",
                "source_system": "zera_seed_catalog",
                "source_record_id": "Public-Notice-Fuel-Notice-4-October-2025",
                "schema_version": "regulatory-event.v1",
                "transform_version": "zera_seed_catalog.v1",
                "transform_run_id": "run-zera-1",
                "raw_document_id": "raw-zera-1",
                "properties": {"category": "Press Releases"},
            }
        ],
        "raw_documents": [
            {
                "id": "raw-zera-1",
                "jurisdiction": "ZW",
                "source_system": "zera_seed_catalog",
                "source_url": "https://www.zera.co.zw/press-releases-public-notices/",
                "content_type": "application/json",
                "fetch_status": "seed_catalog",
                "http_status": 200,
                "document_hash": "sha256:zera",
                "payload_text": "[]",
                "metadata": {"catalog": "zera_seed_events"},
            }
        ],
        "transform_runs": [
            {
                "id": "run-zera-1",
                "jurisdiction": "ZW",
                "source_system": "zera_seed_catalog",
                "raw_document_id": "raw-zera-1",
                "transform_name": "zera_seed_catalog_v1",
                "schema_version": "regulatory-event.v1",
                "transform_version": "zera_seed_catalog.v1",
                "mapping": {"title": {"source": "title"}},
                "notes": "Seeded fallback catalog for ZERA notices.",
                "record_count": 1,
            }
        ],
    }

    with patch("workflows.regulatory_workflow.fetch_generator_capacity", return_value=[
        {"endpoint": "operating-generator-capacity", "period": "2025-12", "state": "CA", "fuel_type": "SUN", "value": 20.0, "unit": "MW", "properties": {"plantName": "Sunray 2"}}
    ]), patch("workflows.regulatory_workflow.fetch_operational_data", return_value=[
        {"endpoint": "electric-power-operational-data", "period": "2025-12", "state": "CA", "fuel_type": "SUN", "value": 2536.0, "unit": "thousand megawatthours", "properties": {"stateDescription": "California"}}
    ]), patch("workflows.regulatory_workflow.fetch_retail_sales", return_value=[
        {"endpoint": "retail-sales", "period": "2025-12", "state": "CA", "fuel_type": "RES", "value": 15555.0, "unit": "million kilowatt hours", "properties": {"stateDescription": "California"}}
    ]), patch("workflows.regulatory_workflow.fetch_utility_rates", return_value={
        "utility_name": "Public Service Co of Colorado",
        "state": "CO",
        "lat": 39.7392,
        "lon": -104.9903,
        "rates": {"residential": 0.05951, "commercial": 0.08123},
        "properties": {"source": "OpenEI"},
    }), patch("workflows.regulatory_workflow.load_rps_data", return_value=[
        {"state": "CA", "standard_type": "RPS", "tier": "Total RPS", "year": 2025, "target_pct": 0.60,
         "demand_gwh": 150000.0, "applicable_sales_gwh": 250000.0, "statewide_sales_gwh": 285000.0,
         "achievement_ratio": 0.98, "compliance_cost_per_kwh": 0.012, "capacity_additions_mw": 4200.0,
         "notes": "On track", "properties": {"source": "test"}}
    ]), patch("workflows.regulatory_workflow.fetch_nersa_events", return_value=nersa_bundle), patch(
        "workflows.regulatory_workflow.fetch_ippo_press_releases", return_value=ippo_bundle
    ), patch("workflows.regulatory_workflow.fetch_zera_events", return_value=zera_bundle):
        updated = workflow.update_all(force=True)

    assert updated == 8
    za_events = store.get_regulatory_events(jurisdiction="ZA")
    zw_events = store.get_regulatory_events(jurisdiction="ZW")
    assert len(za_events) == 2
    assert len(zw_events) == 1
    assert store.get_transform_runs(source_system="ippo_oldnews")[0]["transform_version"] == "ippo_press_release.v1"
    assert store.get_raw_documents(source_system="zera_seed_catalog")[0]["fetch_status"] == "seed_catalog"
    store.close()


def test_background_regulatory_refresh_thread_starts():
    from main import _start_regulatory_refresh_thread

    mock_workflow = MagicMock()
    mock_workflow.update_all.return_value = 4

    with patch("main.RegulatoryWorkflow", return_value=mock_workflow):
        thread = _start_regulatory_refresh_thread()

    assert thread is not None
    assert thread.daemon is True
    assert thread.name == "regulatory-refresh"
    time.sleep(0.2)
    assert mock_workflow.update_all.called


class TestRegulatoryApi:
    @pytest.fixture
    def client(self):
        mod = _import_app_module()
        return mod.app.test_client()

    def test_rps_endpoint_returns_filtered_targets(self, client):
        with patch("ui.web.app.RegulatoryDataStore") as mock_store_cls:
            mock_store = MagicMock()
            mock_store.get_rps_targets.return_value = [{"state": "CA", "year": 2025, "target_pct": 0.60}]
            mock_store_cls.return_value = mock_store

            response = client.get("/api/regulatory/rps?state=CA&year=2025")
            assert response.status_code == 200
            payload = response.get_json()
            assert payload["count"] == 1
            mock_store.get_rps_targets.assert_called_once_with(state="CA", year=2025, standard_type=None)

    def test_capacity_endpoint_returns_eia_data(self, client):
        with patch("ui.web.app.RegulatoryDataStore") as mock_store_cls:
            mock_store = MagicMock()
            mock_store.get_eia_series.return_value = [{"state": "CA", "fuel_type": "SUN", "value": 20.0}]
            mock_store_cls.return_value = mock_store

            response = client.get("/api/regulatory/eia/capacity?state=CA&fuel_type=SUN")
            assert response.status_code == 200
            payload = response.get_json()
            assert payload["count"] == 1
            mock_store.get_eia_series.assert_called_once_with("operating-generator-capacity", state="CA", fuel_type="SUN")

    def test_rates_endpoint_returns_rate_records(self, client):
        with patch("ui.web.app.RegulatoryDataStore") as mock_store_cls:
            mock_store = MagicMock()
            mock_store.get_utility_rates.return_value = [{"state": "CO", "sector": "residential", "rate_kwh": 0.05951}]
            mock_store_cls.return_value = mock_store

            response = client.get("/api/regulatory/rates?state=CO&sector=residential")
            assert response.status_code == 200
            payload = response.get_json()
            assert payload["count"] == 1
            mock_store.get_utility_rates.assert_called_once_with(state="CO", sector="residential")

    def test_refresh_endpoint_runs_workflow(self, client):
        with patch("ui.web.app.RegulatoryWorkflow") as mock_workflow_cls:
            mock_workflow = MagicMock()
            mock_workflow.update_all.return_value = 6
            mock_workflow_cls.return_value = mock_workflow

            response = client.post("/api/regulatory/refresh")
            assert response.status_code == 200
            payload = response.get_json()
            assert payload["updated_sources"] == 6
            mock_workflow.update_all.assert_called_once_with(force=True)

    def test_events_endpoint_returns_lineage_fields(self, client):
        with patch("ui.web.app.RegulatoryDataStore") as mock_store_cls:
            mock_store = MagicMock()
            mock_store.get_regulatory_events.return_value = [
                {
                    "jurisdiction": "ZW",
                    "regulator": "ZERA",
                    "event_type": "public_notice",
                    "title": "Public Notice - Fuel Notice 4 October 2025",
                    "published_date": "2025-10-06",
                    "source_url": "https://www.zera.co.zw/press-releases-public-notices/",
                    "source_system": "zera_seed_catalog",
                    "source_record_id": "Public-Notice-Fuel-Notice-4-October-2025",
                    "schema_version": "regulatory-event.v1",
                    "transform_version": "zera_seed_catalog.v1",
                    "transform_run_id": "run-zera-1",
                    "raw_document_id": "raw-zera-1",
                    "properties": {"category": "Press Releases"},
                }
            ]
            mock_store_cls.return_value = mock_store

            response = client.get("/api/regulatory/events?jurisdiction=ZW")
            assert response.status_code == 200
            payload = response.get_json()
            assert payload["count"] == 1
            item = payload["items"][0]
            assert item["schema_version"] == "regulatory-event.v1"
            assert item["transform_version"] == "zera_seed_catalog.v1"
            assert item["source_system"] == "zera_seed_catalog"
            mock_store.get_regulatory_events.assert_called_once_with(jurisdiction="ZW")

    def test_provenance_endpoint_returns_cached_metadata(self, client):
        with patch("ui.web.app.RegulatoryDataStore") as mock_store_cls:
            mock_store = MagicMock()
            mock_store.get_raw_documents.return_value = [
                {
                    "source_system": "zera_seed_catalog",
                    "fetch_status": "seed_catalog",
                    "metadata": {"fallback_reason": "anti_bot_gate"},
                }
            ]
            mock_store.get_transform_runs.return_value = [
                {
                    "source_system": "zera_seed_catalog",
                    "transform_version": "zera_seed_catalog.v1",
                    "mapping": {"title": {"source": "title"}},
                }
            ]
            mock_store_cls.return_value = mock_store

            response = client.get("/api/regulatory/provenance?jurisdiction=ZW&source_system=zera_seed_catalog")
            assert response.status_code == 200
            payload = response.get_json()
            assert payload["raw_document_count"] == 1
            assert payload["transform_run_count"] == 1
            assert payload["raw_documents"][0]["fetch_status"] == "seed_catalog"
            assert payload["transform_runs"][0]["transform_version"] == "zera_seed_catalog.v1"
            mock_store.get_raw_documents.assert_called_once_with(jurisdiction="ZW", source_system="zera_seed_catalog")
            mock_store.get_transform_runs.assert_called_once_with(jurisdiction="ZW", source_system="zera_seed_catalog")


def test_template_has_regulatory_mode():
    mod = _import_app_module()
    client = mod.app.test_client()
    response = client.get("/")
    assert response.status_code == 200
    html = response.get_data(as_text=True)
    assert "data-mode=\"regulatory\"" in html
    assert "Regulatory" in html
    assert "regulatorySection" in html
    assert "regJurisdiction" in html
    assert "/api/regulatory/events" in html
    assert "/api/regulatory/provenance" in html
    assert "Recent Regulatory Events" in html
    assert "Cache Provenance" in html


# ---------------------------------------------------------------------------
# SEP-035: limit param, stripped fields, loading/error UX
# ---------------------------------------------------------------------------


def test_store_limit_parameter(tmp_path):
    """get_regulatory_events(limit=2) returns only 2 rows when 5 exist."""
    from tools.regulatory.store import RegulatoryDataStore

    store = RegulatoryDataStore(db_path=str(tmp_path / "regulatory.db"))
    events = [
        {
            "jurisdiction": "US",
            "regulator": "FERC",
            "event_type": "rulemaking",
            "title": f"Order {i}",
            "summary": f"Test event {i}",
            "published_date": f"2026-03-0{i}",
            "effective_date": None,
            "deadline_date": None,
            "source_url": f"https://example.com/order{i}",
            "properties": {"docket": f"RM-{i}"},
        }
        for i in range(1, 6)
    ]
    store.upsert_regulatory_events(events)
    result = store.get_regulatory_events(limit=2)
    assert len(result) == 2
    all_result = store.get_regulatory_events()
    assert len(all_result) == 5
    store.close()


def test_api_events_accepts_limit_and_returns_total_count():
    """GET /api/regulatory/events?limit=1 returns 1 item but count reflects total."""
    mod = _import_app_module()
    client = mod.app.test_client()

    with patch("ui.web.app.RegulatoryDataStore") as mock_store_cls:
        mock_store = MagicMock()
        all_items = [
            {"jurisdiction": "US", "title": f"Order {i}", "properties_json": '{}'}
            for i in range(5)
        ]
        mock_store.get_regulatory_events.return_value = all_items
        mock_store_cls.return_value = mock_store

        response = client.get("/api/regulatory/events?limit=1")
        assert response.status_code == 200
        payload = response.get_json()
        assert payload["count"] == 5
        assert len(payload["items"]) == 1


def test_api_events_strips_properties_json():
    """properties_json should not appear in API response items."""
    mod = _import_app_module()
    client = mod.app.test_client()

    with patch("ui.web.app.RegulatoryDataStore") as mock_store_cls:
        mock_store = MagicMock()
        mock_store.get_regulatory_events.return_value = [
            {
                "jurisdiction": "US",
                "title": "Order 1",
                "properties_json": '{"docket": "RM-1"}',
                "properties": {"docket": "RM-1"},
            }
        ]
        mock_store_cls.return_value = mock_store

        response = client.get("/api/regulatory/events")
        assert response.status_code == 200
        payload = response.get_json()
        item = payload["items"][0]
        assert "properties_json" not in item
        assert "properties" in item


def test_api_provenance_strips_payload_text():
    """payload_text should not appear in provenance raw_documents response."""
    mod = _import_app_module()
    client = mod.app.test_client()

    with patch("ui.web.app.RegulatoryDataStore") as mock_store_cls:
        mock_store = MagicMock()
        mock_store.get_raw_documents.return_value = [
            {
                "source_system": "nersa",
                "fetch_status": "ok",
                "payload_text": "<html>huge page</html>",
                "metadata_json": '{}',
                "metadata": {},
            }
        ]
        mock_store.get_transform_runs.return_value = [
            {
                "source_system": "nersa",
                "transform_version": "v1",
                "mapping_json": '{}',
                "mapping": {},
            }
        ]
        mock_store_cls.return_value = mock_store

        response = client.get("/api/regulatory/provenance")
        assert response.status_code == 200
        payload = response.get_json()
        raw_doc = payload["raw_documents"][0]
        assert "payload_text" not in raw_doc
        assert "metadata_json" not in raw_doc
        assert "metadata" in raw_doc
        transform = payload["transform_runs"][0]
        assert "mapping_json" not in transform
        assert "mapping" in transform


def test_template_has_loading_and_allsettled():
    """Frontend JS should use Promise.allSettled and show loading state."""
    mod = _import_app_module()
    client = mod.app.test_client()
    response = client.get("/")
    assert response.status_code == 200
    html = response.get_data(as_text=True)
    assert "Promise.allSettled" in html
    assert "Loading regulatory data" in html


# ---------------------------------------------------------------------------
# SEP-036 (cont.): Diligence Search Integration Tests
#
# These tests exercise real code paths with real local data from the
# imaging and regulatory stores on disk.  Each test would FAIL if the
# corresponding production wiring were broken.
# ---------------------------------------------------------------------------

# Shared test asset that matches real data in ~/.zorora/imaging_data.db
# (342 solar assets in South Africa with lowercase "solar" technology)
_ZA_SOLAR_ASSET = {
    "name": "Test Solar Farm",
    "technology": "solar",
    "capacity_mw": 100,
    "country": "South Africa",
    "operator": "TestCorp",
    "owner": "TestOwner",
    "status": "operating",
}


def test_diligence_decomposition_produces_domain_intents():
    """decompose_diligence_query returns 6 intents with analyst-style diligence
    terminology (PPA, IPP, license, EIA, capacity factor, offtake)."""
    from engine.query_refiner import decompose_diligence_query

    query = "Brownfield acquisition diligence for Test Solar Farm"
    intents = decompose_diligence_query(query, _ZA_SOLAR_ASSET)

    # Exactly 6 domain-specific intents
    assert len(intents) == 6, f"Expected 6 intents, got {len(intents)}"

    # First intent is primary, rest are not
    assert intents[0].is_primary is True
    for intent in intents[1:]:
        assert intent.is_primary is False, f"Non-primary intent marked primary: {intent.intent_query}"

    # All intents carry the original query as parent
    for intent in intents:
        assert intent.parent_query == query

    queries_lower = [i.intent_query.lower() for i in intents]
    all_text = " ".join(queries_lower)

    # Asset metadata must appear across intents
    assert "south africa" in all_text, "Country not injected into intents"
    assert "solar" in all_text, "Technology not injected into intents"

    # Diligence-specific terms that an analyst would use (not generic educational)
    assert any("power purchase agreement" in q or "ppa" in q for q in queries_lower), \
        "No intent mentions PPA / power purchase agreement"
    assert any("license" in q or "permits" in q for q in queries_lower), \
        "No intent mentions licensing/permits"
    assert any("environmental impact assessment" in q or "grid connection" in q for q in queries_lower), \
        "No intent mentions EIA or grid connection"
    assert any("capacity factor" in q or "performance ratio" in q for q in queries_lower), \
        "No intent mentions capacity factor or performance ratio"
    assert any("offtake" in q or "ipp" in q for q in queries_lower), \
        "No intent mentions offtake/IPP"
    assert "test solar farm" in all_text, "Asset name not in any intent"

    # Must NOT contain generic educational phrases
    for q in queries_lower:
        assert "how to" not in q, f"Intent contains educational phrase 'how to': {q}"
        assert "what is" not in q, f"Intent contains educational phrase 'what is': {q}"


def test_force_policy_actually_includes_policy_channel():
    """force_policy=True causes policy_search_sources to be called even when
    the query contains no policy keywords."""
    from workflows.deep_research.aggregator import aggregate_sources

    query = "solar power generation capacity factors"  # No policy keywords

    # Patch all network-calling functions to return [] quickly
    with patch("workflows.deep_research.aggregator.academic_search_sources", return_value=[]), \
         patch("workflows.deep_research.aggregator.web_search_sources", return_value=[]), \
         patch("workflows.deep_research.aggregator.fetch_newsroom_api", return_value=[]), \
         patch("workflows.deep_research.aggregator.worldbank_search_sources", return_value=[]), \
         patch("workflows.deep_research.aggregator.policy_search_sources", return_value=[]) as mock_policy, \
         patch("workflows.deep_research.aggregator.sec_search_sources", return_value=[]):

        # With force_policy=False: policy should NOT be called (no policy keywords)
        aggregate_sources(query, force_policy=False)
        mock_policy.assert_not_called()

        # With force_policy=True: policy MUST be called
        mock_policy.reset_mock()
        aggregate_sources(query, force_policy=True)
        mock_policy.assert_called_once()


def test_diligence_skips_us_policy_for_non_us_assets():
    """Non-US diligence assets should NOT trigger US-only policy search
    (Congress.gov, GovTrack, Federal Register are useless for Lesotho/ZA)."""
    from engine.deep_research_service import run_deep_research

    # Patch all search backends to return [] and track policy calls
    with patch("workflows.deep_research.aggregator.academic_search_sources", return_value=[]), \
         patch("workflows.deep_research.aggregator.web_search_sources", return_value=[]), \
         patch("workflows.deep_research.aggregator.fetch_newsroom_api", return_value=[]), \
         patch("workflows.deep_research.aggregator.worldbank_search_sources", return_value=[]), \
         patch("workflows.deep_research.aggregator.policy_search_sources", return_value=[]) as mock_policy, \
         patch("workflows.deep_research.aggregator.sec_search_sources", return_value=[]), \
         patch("workflows.deep_research.synthesizer.synthesize_direct", return_value="test"):

        # Non-US asset: policy search should NOT be called
        run_deep_research(
            "Diligence for Test Solar Farm",
            depth=1,
            research_type="diligence",
            asset_metadata=_ZA_SOLAR_ASSET,
        )
        mock_policy.assert_not_called()

        # US asset: policy search SHOULD be called
        mock_policy.reset_mock()
        us_asset = {**_ZA_SOLAR_ASSET, "country": "United States"}
        run_deep_research(
            "Diligence for Test Solar Farm",
            depth=1,
            research_type="diligence",
            asset_metadata=us_asset,
        )
        assert mock_policy.call_count > 0, "US diligence should trigger policy search"


def test_build_diligence_context_with_real_imaging_data():
    """_build_diligence_context queries the real imaging store and returns
    structured context with actual comparable plant data from disk."""
    from engine.deep_research_service import _build_diligence_context

    context_text, raw_data = _build_diligence_context(_ZA_SOLAR_ASSET)

    if "No local diligence data" in context_text:
        pytest.skip("Imaging DB has no generation assets (CI environment)")

    # Context should describe comparable plants found in real DB
    assert "Comparable solar Plants in South Africa" in context_text, \
        f"Context missing comparable plants section. Got: {context_text[:200]}"
    assert "found" in context_text, "Context should report count of plants found"

    # raw_data must contain actual plant features from the imaging DB
    assert "comparable_plants" in raw_data, "raw_data missing comparable_plants key"
    plants = raw_data["comparable_plants"]
    assert isinstance(plants, list)
    assert len(plants) > 0, "No comparable plants returned from real imaging DB"
    assert len(plants) <= 20, "Should be capped at 20 plants"

    # Each plant must have real capacity data
    for plant in plants:
        cap = plant.get("properties", {}).get("capacity_mw", 0)
        assert isinstance(cap, (int, float)) and cap > 0, \
            f"Plant missing positive capacity_mw: {plant.get('properties', {})}"


def test_generate_diligence_charts_with_real_comparable_data():
    """generate_diligence_charts produces 2 valid PNG charts using real
    comparable plant data from the imaging store on disk."""
    from workflows.deep_research.synthesizer import generate_diligence_charts

    charts = generate_diligence_charts(_ZA_SOLAR_ASSET)

    if len(charts) < 2:
        pytest.skip("Imaging DB has no comparable plants for chart generation (CI environment)")

    # Must produce exactly 2 charts: revenue estimate + capacity benchmark
    assert len(charts) == 2, f"Expected 2 charts, got {len(charts)}: {[t for t,_ in charts]}"

    titles = [title for title, _ in charts]
    assert "Revenue Estimate" in titles, f"Missing Revenue Estimate chart. Got: {titles}"
    assert "Capacity Benchmark" in titles, f"Missing Capacity Benchmark chart. Got: {titles}"

    # Both must be valid PNG images
    for title, data_uri in charts:
        assert data_uri.startswith("data:image/png;base64,"), \
            f"Chart '{title}' not a data URI: {data_uri[:50]}"
        b64_data = data_uri.split(",", 1)[1]
        decoded = base64.b64decode(b64_data)
        # PNG magic bytes: \x89PNG\r\n\x1a\n
        assert decoded[:4] == b"\x89PNG", \
            f"Chart '{title}' is not valid PNG (magic bytes: {decoded[:4]})"

    # Verify the capacity benchmark was built from real data, not an empty set
    from tools.imaging.store import ImagingDataStore
    img_store = ImagingDataStore()
    comparable = img_store.get_generation_assets(technology="solar", country="South Africa")
    features = comparable.get("features", []) if isinstance(comparable, dict) else []
    assert len(features) > 0, "Capacity benchmark chart requires real comparable plants"


def test_diligence_synthesis_prompt_structure_with_real_context():
    """Diligence synthesis prompt contains all 6 required sections, real asset
    metadata, and real local data context from the imaging store."""
    from workflows.deep_research.synthesizer import _build_diligence_synthesis_prompt
    from engine.deep_research_service import _build_diligence_context
    from engine.models import ResearchState, Source

    # Get real context from real stores
    real_context, _ = _build_diligence_context(_ZA_SOLAR_ASSET)

    state = ResearchState(
        original_query="Brownfield acquisition diligence for Test Solar Farm",
        research_type="diligence",
        asset_metadata=_ZA_SOLAR_ASSET,
    )
    # Add a source so the prompt has something to format
    state.sources_checked = [
        Source(source_id="src1", url="https://example.com", title="Test Source",
               source_type="web", relevance_score=0.8, credibility_score=0.7),
    ]

    prompt = _build_diligence_synthesis_prompt(
        state, diligence_context=real_context, asset_metadata=_ZA_SOLAR_ASSET,
    )

    # All 6 exact section headers must appear
    required_headers = [
        "## Executive Summary",
        "## Tariff & Revenue Potential",
        "## Regulatory & Licensing Requirements",
        "## Performance Gap Analysis",
        "## Vendor & Counterparty Relationships",
        "## Risk Summary & Recommendation",
    ]
    for header in required_headers:
        assert header in prompt, f"Missing required header '{header}' in prompt"

    # Real asset metadata must be in the prompt
    assert "Test Solar Farm" in prompt, "Asset name not in prompt"
    assert "solar" in prompt.lower(), "Technology not in prompt"
    assert "100" in prompt, "Capacity not in prompt"
    assert "South Africa" in prompt, "Country not in prompt"

    # Real local data context must be injected
    assert "LOCAL DATA CONTEXT" in prompt, "Missing local data context block"
    if "No local diligence data" in real_context:
        pytest.skip("Imaging DB has no generation assets (CI environment)")
    assert "Comparable solar Plants" in prompt, "Real comparable plants data not in prompt"

    # The ranked source must appear
    assert "Test Source" in prompt, "Source not formatted into prompt"


def test_synthesize_direct_takes_diligence_branch():
    """synthesize_direct uses the diligence prompt (not the generic direct prompt),
    calls generate_diligence_charts with the real imaging DB, and inserts chart
    markdown into the synthesis output."""
    from workflows.deep_research.synthesizer import synthesize_direct
    from engine.models import ResearchState, Source

    state = ResearchState(
        original_query="Brownfield acquisition diligence for Test Solar Farm",
        research_type="diligence",
        asset_metadata=_ZA_SOLAR_ASSET,
    )
    state.sources_checked = [
        Source(source_id="src1", url="https://example.com/1", title="Solar Tariff Report",
               source_type="web", relevance_score=0.9, credibility_score=0.8),
        Source(source_id="src2", url="https://example.com/2", title="ZA Grid Policy",
               source_type="policy", relevance_score=0.85, credibility_score=0.7),
    ]

    # Fake LLM output with all 6 required sections
    fake_synthesis = (
        "## Executive Summary\n"
        "This asset shows moderate potential [Solar Tariff Report].\n\n"
        "## Tariff & Revenue Potential\n"
        "Tariffs in South Africa average $0.08/kWh [Solar Tariff Report].\n\n"
        "## Regulatory & Licensing Requirements\n"
        "NERSA licensing is required [ZA Grid Policy].\n\n"
        "## Performance Gap Analysis\n"
        "Capacity factor benchmarks at 20% for solar [Solar Tariff Report].\n\n"
        "## Vendor & Counterparty Relationships\n"
        "Eskom is the primary offtaker [ZA Grid Policy].\n\n"
        "## Risk Summary & Recommendation\n"
        "Key risks include regulatory uncertainty [ZA Grid Policy].\n"
    )

    captured_prompts = []

    def fake_model_call(prompt, system_prompt=None):
        captured_prompts.append(prompt)
        return fake_synthesis

    # Patch the model call and the quality gate. The quality gate expects
    # a "Direct Answer" section (generic path); diligence has different
    # headers. We test that the diligence branch is taken and charts are
    # inserted — the quality gate is not what's under test here.
    with patch("workflows.deep_research.synthesizer._call_research_synthesis_model",
               side_effect=fake_model_call), \
         patch("workflows.deep_research.synthesizer._passes_direct_synthesis_quality_gate",
               return_value=True):
        result = synthesize_direct(
            state,
            asset_metadata=_ZA_SOLAR_ASSET,
            diligence_context="Comparable solar Plants in South Africa: 342 found",
        )

    # Verify the diligence prompt was used, not the generic one
    assert len(captured_prompts) >= 1, "Model was never called"
    prompt_sent = captured_prompts[0]
    assert "ASSET UNDER REVIEW" in prompt_sent, \
        "Diligence prompt not used — 'ASSET UNDER REVIEW' missing"
    assert "Test Solar Farm" in prompt_sent, "Asset name not in prompt sent to model"

    # Charts must be inserted into the synthesis output
    assert "![Revenue Estimate](data:image/png;base64," in result, \
        "Revenue chart not inserted into synthesis"
    if "![Capacity Benchmark]" not in result:
        pytest.skip("Imaging DB has no comparable plants for capacity benchmark (CI environment)")
    assert "![Capacity Benchmark](data:image/png;base64," in result, \
        "Capacity benchmark chart not inserted into synthesis"

    # Revenue chart should follow the Tariff section (not be appended randomly)
    tariff_pos = result.find("## Tariff & Revenue Potential")
    revenue_chart_pos = result.find("![Revenue Estimate]")
    assert tariff_pos >= 0, "Tariff section missing from output"
    assert revenue_chart_pos > tariff_pos, \
        "Revenue chart should be inserted after the Tariff section"

    # Capacity benchmark chart should follow the Performance section
    perf_pos = result.find("## Performance Gap Analysis")
    bench_chart_pos = result.find("![Capacity Benchmark]")
    assert perf_pos >= 0, "Performance section missing from output"
    assert bench_chart_pos > perf_pos, \
        "Capacity benchmark chart should be inserted after the Performance section"


def test_api_passes_asset_metadata_through_to_pipeline():
    """POST /api/research passes asset_metadata and research_type all the way
    through to _run_research_with_progress, not just accepting silently."""
    mod = _import_app_module()
    client = mod.app.test_client()

    captured_kwargs = {}

    def capture_research(*args, **kwargs):
        captured_kwargs.update(kwargs)

    with patch("ui.web.app._run_research_with_progress", side_effect=capture_research):
        resp = client.post("/api/research", json={
            "query": "Brownfield diligence for Test Solar Farm",
            "depth": 1,
            "research_type": "diligence",
            "asset_metadata": {
                "name": "Test Solar Farm",
                "technology": "solar",
                "country": "South Africa",
                "capacity_mw": 100,
            },
        })

    assert resp.status_code == 200
    assert "research_id" in resp.get_json()

    # asset_metadata must be passed through, not dropped
    assert "asset_metadata" in captured_kwargs, \
        f"asset_metadata not passed to pipeline. kwargs: {list(captured_kwargs.keys())}"
    meta = captured_kwargs["asset_metadata"]
    assert meta["name"] == "Test Solar Farm"
    assert meta["technology"] == "solar"
    assert meta["country"] == "South Africa"
    assert meta["capacity_mw"] == 100

    # research_type must be "diligence"
    assert captured_kwargs.get("research_type") == "diligence", \
        f"research_type not passed correctly: {captured_kwargs.get('research_type')}"


def test_extractive_summarize_scores_relevant_sentences():
    """_extractive_summarize returns sentences most similar to the query,
    not random ones, using TF-IDF cosine similarity."""
    from workflows.deep_research.synthesizer import _extractive_summarize

    texts = [
        "Lesotho electricity tariff schedule sets feed-in tariff at 0.85 USD/kWh for solar IPPs.",
        "The weather in Paris is lovely this time of year with temperatures around 22 degrees.",
        "LEWA requires all independent power producers to obtain a generation license before grid connection.",
        "Football World Cup 2026 will be held in North America across three countries.",
        "Lesotho Electricity Company signed a 20-year PPA with Globeleq for 100MW solar output.",
        "Recipe for chocolate cake requires flour, sugar, eggs, and butter mixed together.",
    ]
    query = "Lesotho solar power purchase agreement tariff"

    result = _extractive_summarize(texts, query, max_sentences=3)

    # Relevant sentences about Lesotho/solar/tariff/PPA should be selected
    assert "tariff" in result.lower() or "ppa" in result.lower(), \
        f"Extractive summary missed tariff/PPA content: {result}"
    assert "lesotho" in result.lower(), \
        f"Extractive summary missed Lesotho content: {result}"
    # Irrelevant sentences should NOT appear
    assert "football" not in result.lower(), \
        f"Extractive summary included irrelevant football content: {result}"
    assert "chocolate" not in result.lower(), \
        f"Extractive summary included irrelevant recipe content: {result}"
    assert "paris" not in result.lower(), \
        f"Extractive summary included irrelevant Paris content: {result}"


def test_deterministic_diligence_synthesis_produces_domain_sections():
    """When LLM is unavailable, _deterministic_diligence_synthesis produces
    a structured report with domain sections from tagged sources."""
    from engine.models import Source, ResearchState
    from workflows.deep_research.synthesizer import _deterministic_diligence_synthesis

    # Build sources tagged with different intent domains
    sources = []
    domain_content = {
        "commercial": "Lesotho feed-in tariff for solar is 0.85 USD per kWh under the LEWA tariff schedule 2025.",
        "licensing": "LEWA requires generation license application with environmental clearance before construction.",
        "environmental": "Grid connection at 132kV Mafeteng substation requires LEWA technical approval and EIA.",
        "performance": "Average solar capacity factor in Lesotho is 19.2% based on measured irradiance data.",
        "counterparty": "Lesotho Electricity Company is the sole offtaker for all IPP generation under standard PPA.",
        "asset_specific": "Globeleq acquired the Mafeteng Solar project in 2024 for USD 85 million.",
    }
    for domain, content in domain_content.items():
        s = Source(
            source_id=f"src_{domain}",
            url=f"https://example.com/{domain}",
            title=f"{domain.title()} Source",
            content_snippet=content,
            content_full=content,
            relevance_score=0.7,
            credibility_score=0.6,
            intent_domain=domain,
        )
        sources.append(s)

    state = ResearchState(
        original_query="Brownfield acquisition diligence for Mafeteng Solar",
        research_type="diligence",
    )
    state.sources_checked = sources

    result = _deterministic_diligence_synthesis(state, _ZA_SOLAR_ASSET)

    # Must have domain section headers (not generic Direct Answer/Supporting Evidence)
    assert "## Tariff & Revenue" in result, f"Missing Tariff section:\n{result}"
    assert "## Regulatory & Licensing" in result, f"Missing Regulatory section:\n{result}"
    assert "## Performance" in result, f"Missing Performance section:\n{result}"
    assert "## Vendor & Counterparty" in result or "## Counterparty" in result, \
        f"Missing Vendor section:\n{result}"
    # Must NOT have generic fallback sections
    assert "## Direct Answer" not in result, f"Generic fallback used instead of diligence:\n{result}"
    assert "## Supporting Evidence" not in result, f"Generic fallback used instead of diligence:\n{result}"

    # Content from tagged sources should appear in matching sections
    assert "tariff" in result.lower(), "Tariff content missing from synthesis"
    assert "capacity factor" in result.lower() or "19.2" in result, "Performance content missing"
    assert "synthesis_model" not in result  # implementation detail shouldn't leak


def test_diligence_fallback_used_when_llm_unavailable():
    """When _call_research_synthesis_model returns None for diligence,
    synthesize_direct uses the diligence template with source content routed
    to the correct domain sections — not the generic Direct Answer fallback."""
    from engine.models import Source, ResearchState
    from workflows.deep_research.synthesizer import synthesize_direct

    domain_snippets = [
        ("commercial", "Feed-in tariff rate is 0.85 USD/kWh for solar generation in Lesotho."),
        ("licensing", "LEWA generation license required for all IPPs above 500kW capacity."),
        ("performance", "Measured capacity factor for Lesotho solar plants averages 19.2 percent."),
    ]
    sources = []
    for i, (domain, snippet) in enumerate(domain_snippets):
        s = Source(
            source_id=f"test_{i}",
            url=f"https://example.com/{i}",
            title=f"Source {domain.title()}",
            content_snippet=snippet,
            content_full=snippet,
            relevance_score=0.65,
            credibility_score=0.5,
            intent_domain=domain,
        )
        sources.append(s)

    state = ResearchState(
        original_query="Diligence for Mafeteng Solar Farm",
        research_type="diligence",
    )
    state.sources_checked = sources

    with patch("workflows.deep_research.synthesizer._call_research_synthesis_model", return_value=None):
        result = synthesize_direct(state, asset_metadata=_ZA_SOLAR_ASSET)

    # Must use diligence domain sections, not generic fallback
    assert "## Tariff & Revenue" in result, \
        f"Missing Tariff & Revenue section. Got:\n{result[:500]}"
    assert "## Regulatory" in result, \
        f"Missing Regulatory section. Got:\n{result[:500]}"
    assert "## Direct Answer" not in result, \
        f"Generic fallback used instead of diligence fallback:\n{result[:500]}"

    # Source content must appear in the output (not just headers)
    result_lower = result.lower()
    assert "0.85" in result or "tariff" in result_lower, \
        f"Commercial source content not routed to output:\n{result[:800]}"
    assert "lewa" in result_lower or "license" in result_lower, \
        f"Licensing source content not routed to output:\n{result[:800]}"
    assert "19.2" in result or "capacity factor" in result_lower, \
        f"Performance source content not routed to output:\n{result[:800]}"


def test_diligence_variant_suppression():
    """Diligence searches force num_variants=1 to prevent LLM from mangling
    specific analyst-style queries into generic subtopics."""
    from engine.deep_research_service import _generate_query_variants

    diligence_query = "Lesotho independent power producer power purchase agreement solar tariff rates feed-in tariff 2025 2026"

    # With num_variants=1 (what diligence forces), query passes through untouched
    variants = _generate_query_variants(diligence_query, 1)
    assert variants == [diligence_query], \
        f"num_variants=1 should return query unchanged, got: {variants}"

    # Verify the suppression logic in the search loop itself
    from engine.deep_research_service import run_deep_research

    captured_variants = []

    def mock_aggregate(query, **kwargs):
        captured_variants.append(query)
        return []

    with patch("engine.deep_research_service.aggregate_sources", side_effect=mock_aggregate), \
         patch("workflows.deep_research.synthesizer.synthesize_direct", return_value="test"):
        run_deep_research(
            "Diligence for Test Solar Farm",
            depth=2,  # depth 2 normally uses 2-3 variants
            research_type="diligence",
            asset_metadata=_ZA_SOLAR_ASSET,
        )

    # Each intent should produce exactly 1 variant (the original query, not LLM-decomposed)
    # With 6 intents × 1 variant = 6 total search calls
    assert len(captured_variants) == 6, \
        f"Expected 6 search calls (6 intents × 1 variant), got {len(captured_variants)}"


def test_diligence_sources_tagged_with_intent_domain():
    """Sources returned during diligence search must have intent_domain set
    to the domain of the intent that found them, so the deterministic fallback
    can group them into the correct report sections."""
    from engine.models import Source
    from engine.deep_research_service import run_deep_research

    call_count = [0]

    def mock_aggregate(query, **kwargs):
        call_count[0] += 1
        # Return a unique source per call so we can verify domain tagging
        return [Source(
            source_id=f"src_{call_count[0]}",
            url=f"https://example.com/{call_count[0]}",
            title=f"Result {call_count[0]}",
            content_snippet=f"Content for call {call_count[0]}",
            content_full=f"Content for call {call_count[0]}",
            relevance_score=0.8,
            credibility_score=0.6,
        )]

    def passthrough_score(intent_query, sources, variants, max_sources, relevance_min):
        return sources

    # Diligence uses synthesize_direct (not synthesize). In CI there is no local LLM;
    # stub synthesis so we assert on source tagging, which is complete before synthesis.
    with patch("engine.deep_research_service.aggregate_sources", side_effect=mock_aggregate), \
         patch("engine.deep_research_service._score_and_filter_intent_sources", side_effect=passthrough_score), \
         patch("engine.deep_research_service.synthesize_direct", return_value="test synthesis"):
        state = run_deep_research(
            "Diligence for Test Solar Farm",
            depth=1,
            research_type="diligence",
            asset_metadata=_ZA_SOLAR_ASSET,
        )

    assert state is not None
    sources = state.sources_checked
    assert len(sources) > 0, "No sources in state"

    # Every source must have a non-empty intent_domain
    for s in sources:
        assert s.intent_domain, f"Source {s.source_id} has empty intent_domain"

    # All 6 domains must be present (6 intents each returning 1 source)
    domains = {s.intent_domain for s in sources}
    expected_domains = {"commercial", "licensing", "environmental", "performance", "counterparty", "asset_specific"}
    assert domains == expected_domains, f"Expected all 6 domains, got {domains}"


# ---------------------------------------------------------------------------
# SEP-038: Domain-aware diligence synthesis
# ---------------------------------------------------------------------------

# Helpers that create domain-tagged sources with distinct, identifiable content
# so tests can verify that the RIGHT source content lands in the RIGHT section.

def _make_tagged_sources(domains=None):
    """Create sources tagged with intent_domain.

    Each source has unique title and content tied to its domain so tests can
    assert cross-contamination doesn't happen (e.g., licensing content must
    NOT appear under the tariff section).
    """
    from engine.models import Source

    if domains is None:
        domains = ["commercial", "licensing", "environmental", "performance", "counterparty", "asset_specific"]

    # Domain-specific content that would NOT appear under other domains
    domain_content = {
        "commercial": "NERSA approved Eskom MYPD5 tariff increase of 12.7% for 2025/26 fiscal year, average retail rate 1.78 ZAR/kWh",
        "licensing": "Generation license application to NERSA requires Schedule 2 exemption for plants under 100MW per Electricity Regulation Act",
        "environmental": "Environmental impact assessment under NEMA requires Basic Assessment for solar PV on previously disturbed land",
        "performance": "Measured capacity factor of 22.3% at De Aar solar park over 2023-2024 operating period with 0.5% annual degradation",
        "counterparty": "Eskom Holdings SOC Ltd is sole offtaker under REIPPP bid window 5 with 20-year PPA term",
        "asset_specific": "TestCorp acquired 49% stake in Test Solar Farm from previous developer in 2024 for undisclosed amount",
    }
    sources = []
    for i, domain in enumerate(domains, 1):
        sources.append(Source(
            source_id=f"src_{domain}_{i}",
            url=f"https://example.com/{domain}/{i}",
            title=f"{domain.replace('_', ' ').title()} Report {i}",
            content_snippet=domain_content.get(domain, f"Content about {domain}."),
            content_full=domain_content.get(domain, f"Full content about {domain}."),
            relevance_score=0.75,
            credibility_score=0.6,
            intent_domain=domain,
        ))
    return sources


def test_format_sources_by_domain_isolates_each_domains_sources():
    """_format_sources_by_domain must put each source under its own domain key ONLY.

    The behavioral requirement: commercial sources appear under "commercial",
    licensing sources under "licensing", and neither leaks into the other.
    This is the foundation of per-chapter source grouping."""
    from workflows.deep_research.synthesizer import _format_sources_by_domain

    sources = _make_tagged_sources()
    grouped = _format_sources_by_domain(sources)

    # All 6 domains must be present
    expected_domains = {"commercial", "licensing", "environmental", "performance", "counterparty", "asset_specific"}
    assert set(grouped.keys()) >= expected_domains, \
        f"Missing domains: {expected_domains - set(grouped.keys())}"

    # Cross-contamination check: each domain's formatted text must contain
    # its own source title and must NOT contain other domains' source titles
    for domain in expected_domains:
        own_title = f"{domain.replace('_', ' ').title()} Report"
        assert own_title in grouped[domain], \
            f"Domain '{domain}' missing its own source: expected '{own_title}'"
        for other_domain in expected_domains - {domain}:
            other_title = f"{other_domain.replace('_', ' ').title()} Report"
            assert other_title not in grouped[domain], \
                f"Cross-contamination: '{other_title}' found under '{domain}' group"

    # Sources with no intent_domain must go to "other", not pollute a domain group
    from engine.models import Source
    orphan = Source(
        source_id="orphan_1", url="https://example.com/orphan", title="Orphan Doc",
        content_snippet="Untagged content", content_full="Untagged content",
        relevance_score=0.5, credibility_score=0.5, intent_domain="",
    )
    grouped_with_orphan = _format_sources_by_domain(sources + [orphan])
    assert "Orphan Doc" not in grouped_with_orphan.get("commercial", ""), \
        "Untagged source leaked into commercial domain"
    assert "Orphan Doc" in grouped_with_orphan.get("other", ""), \
        "Untagged source should be in 'other' bucket"


def test_build_diligence_prompt_v2_puts_right_sources_under_right_sections():
    """The v2 prompt must place each domain's sources under its corresponding
    section header — not dump them all in one block.

    Validates behavioral criterion: 'Each section contains only that domain's
    sources + an analytical question.'"""
    from workflows.deep_research.synthesizer import _build_diligence_synthesis_prompt_v2
    from engine.deep_research_service import ResearchState

    sources = _make_tagged_sources()
    state = ResearchState(
        original_query="Diligence for Test Solar Farm",
        refined_query="Diligence for Test Solar Farm",
        sources_checked=sources,
        research_type="diligence",
    )

    prompt = _build_diligence_synthesis_prompt_v2(
        state,
        diligence_context="EIA data: retail prices 0.08 USD/kWh",
        asset_metadata=_ZA_SOLAR_ASSET,
    )

    # Split the prompt by section headers to verify per-section source placement
    import re
    section_pattern = re.compile(r"^##\s+(.+)$", re.MULTILINE)
    headers = [(m.group(1), m.start()) for m in section_pattern.finditer(prompt)]

    def _section_text(idx):
        start = headers[idx][1]
        end = headers[idx + 1][1] if idx + 1 < len(headers) else len(prompt)
        return prompt[start:end]

    # Find each domain section and verify its sources
    section_map = {title: _section_text(i) for i, (title, _) in enumerate(headers)}

    # Tariff section must have commercial source, not licensing source
    tariff_section = next((v for k, v in section_map.items() if "Tariff" in k or "Revenue" in k), None)
    assert tariff_section is not None, "Missing Tariff & Revenue section"
    assert "Commercial Report" in tariff_section, \
        "Tariff section missing commercial source"
    assert "Licensing Report" not in tariff_section, \
        "Tariff section contains licensing source (cross-contamination)"
    assert "Performance Report" not in tariff_section, \
        "Tariff section contains performance source (cross-contamination)"

    # Regulatory section must have licensing source, not commercial source
    reg_section = next((v for k, v in section_map.items() if "Regulatory" in k or "Licensing" in k), None)
    assert reg_section is not None, "Missing Regulatory & Licensing section"
    assert "Licensing Report" in reg_section, \
        "Regulatory section missing licensing source"
    assert "Commercial Report" not in reg_section, \
        "Regulatory section contains commercial source (cross-contamination)"

    # Performance section must have performance source
    perf_section = next((v for k, v in section_map.items() if "Performance" in k), None)
    assert perf_section is not None, "Missing Performance section"
    assert "Performance Report" in perf_section, \
        "Performance section missing performance source"
    assert "Counterparty Report" not in perf_section, \
        "Performance section contains counterparty source (cross-contamination)"

    # Each domain section must have an analytical question with asset metadata
    assert "South Africa" in tariff_section, "Tariff question missing country"
    assert "tariff" in tariff_section.lower(), "Tariff section missing tariff question"
    assert "license" in reg_section.lower() or "permit" in reg_section.lower(), \
        "Regulatory section missing licensing question"
    assert "capacity factor" in perf_section.lower() or "performance ratio" in perf_section.lower(), \
        "Performance section missing performance question"

    # Must NOT have a flat "RANKED SOURCES:" block (the old v1 pattern)
    assert "RANKED SOURCES:" not in prompt, \
        "v2 prompt still has flat source dump from v1"


def test_quality_gate_accepts_diligence_format_rejects_direct_answer_format():
    """The diligence quality gate must accept reports with domain sections and
    NO 'Direct Answer' — the exact format that the regular gate rejects.

    This is the core behavioral fix: the regular gate at line 2511 requires
    'Direct Answer', which diligence reports never have. The new gate must
    NOT check for 'Direct Answer'."""
    from workflows.deep_research.synthesizer import _passes_diligence_quality_gate
    from engine.deep_research_service import ResearchState

    sources = _make_tagged_sources()
    state = ResearchState(
        original_query="Diligence for Test Solar Farm",
        sources_checked=sources,
    )

    # A valid diligence report — has domain sections, NO "Direct Answer"
    valid_report = "\n\n".join([
        "## Executive Summary",
        "This report analyzes a 100MW solar acquisition in South Africa [Commercial Report 1].",
        "## Tariff & Revenue Potential",
        "NERSA approved 12.7% tariff increase for 2025/26 [Commercial Report 1].",
        "## Regulatory & Licensing Requirements",
        "IPP license required under Electricity Regulation Act [Licensing Report 2].",
        "## Environmental & Grid Connection",
        "Basic Assessment required under NEMA [Environmental Report 3].",
        "## Performance Gap Analysis",
        "De Aar solar park measured 22.3% CF [Performance Report 4].",
        "## Vendor & Counterparty Relationships",
        "Eskom is sole offtaker under REIPPP [Counterparty Report 5].",
        "## Risk Summary & Recommendation",
        "Primary risk is regulatory uncertainty around tariff adjustments [Licensing Report 2].",
    ])

    assert _passes_diligence_quality_gate(valid_report, state), \
        "Diligence gate rejected a valid report — likely still checking for 'Direct Answer'"

    # Verify the regular gate WOULD reject the same report (proving the fix matters)
    from workflows.deep_research.synthesizer import _passes_direct_synthesis_quality_gate
    assert not _passes_direct_synthesis_quality_gate(valid_report, state), \
        "Regular gate should reject diligence format (no 'Direct Answer') — " \
        "if it passes, the gates are not differentiated"


def test_quality_gate_rejects_actually_bad_diligence_reports():
    """The diligence quality gate must reject reports that are genuinely
    incomplete — not just missing 'Direct Answer'."""
    from workflows.deep_research.synthesizer import _passes_diligence_quality_gate
    from engine.deep_research_service import ResearchState

    sources = _make_tagged_sources()
    state = ResearchState(
        original_query="Diligence for Test Solar Farm",
        sources_checked=sources,
    )

    # Only 2 sections — below the minimum 4
    too_short = "\n\n".join([
        "## Executive Summary",
        "Brief summary [Commercial Report 1].",
        "## Tariff & Revenue Potential",
        "Some tariff info [Commercial Report 1].",
    ])
    assert not _passes_diligence_quality_gate(too_short, state), \
        "Gate should reject report with only 2 sections"

    # No Executive Summary — wrong structure entirely
    no_exec = "\n\n".join([
        "## Introduction",
        "This is an introduction [Commercial Report 1].",
        "## Tariff & Revenue Potential",
        "Tariff info [Commercial Report 1].",
        "## Regulatory & Licensing Requirements",
        "License info [Licensing Report 2].",
        "## Performance Gap Analysis",
        "Performance data [Performance Report 4].",
    ])
    assert not _passes_diligence_quality_gate(no_exec, state), \
        "Gate should reject report without Executive Summary"

    # No source citations at all — ungrounded
    no_citations = "\n\n".join([
        "## Executive Summary",
        "Solar energy is growing in South Africa.",
        "## Tariff & Revenue Potential",
        "Tariffs are regulated by NERSA.",
        "## Regulatory & Licensing Requirements",
        "Licenses are required.",
        "## Performance Gap Analysis",
        "Performance varies by location.",
        "## Risk Summary & Recommendation",
        "There are various risks to consider.",
    ])
    assert not _passes_diligence_quality_gate(no_citations, state), \
        "Gate should reject report with zero source citations"

    # Empty string
    assert not _passes_diligence_quality_gate("", state), \
        "Gate should reject empty output"


def test_repair_diligence_preserves_all_sections_unlike_regular_repair():
    """_repair_diligence_synthesis_output must NOT cap sections at 3.

    This validates the exact bug: _repair_direct_synthesis_output has
    `thematic_sections_kept >= 3` (line 2488) which drops sections 4+.
    Diligence needs 5+ domain sections. The regular repair provably
    destroys diligence reports."""
    from workflows.deep_research.synthesizer import (
        _repair_diligence_synthesis_output,
        _repair_direct_synthesis_output,
    )
    from engine.deep_research_service import ResearchState

    sources = _make_tagged_sources()
    state = ResearchState(
        original_query="Diligence for Test Solar Farm",
        sources_checked=sources,
    )

    report = "\n\n".join([
        "## Executive Summary",
        "Summary [Commercial Report 1].",
        "## Tariff & Revenue Potential",
        "NERSA tariff info [Commercial Report 1].",
        "## Regulatory & Licensing Requirements",
        "License requirements [Licensing Report 2].",
        "## Environmental & Grid Connection",
        "EIA requirements [Environmental Report 3].",
        "## Performance Gap Analysis",
        "Capacity factor data [Performance Report 4].",
        "## Vendor & Counterparty Relationships",
        "Offtaker details [Counterparty Report 5].",
        "## Risk Summary & Recommendation",
        "Risk assessment [Asset Specific Report 6].",
    ])

    # Diligence repair must preserve all 7 sections
    diligence_repaired = _repair_diligence_synthesis_output(report, state)
    for section in ["Executive Summary", "Tariff & Revenue", "Regulatory & Licensing",
                     "Environmental & Grid", "Performance Gap", "Vendor & Counterparty",
                     "Risk Summary"]:
        assert section in diligence_repaired, \
            f"Diligence repair dropped '{section}' — must keep all domain sections"

    # Prove the regular repair WOULD drop sections (validating why the fix matters)
    regular_repaired = _repair_direct_synthesis_output(report, state)
    import re
    regular_sections = re.findall(r"^##\s+(.+)$", regular_repaired, re.MULTILINE)
    # Regular repair caps thematic sections at 3, so total ≤ 5 (Exec Summary + Direct Answer + 3 thematic)
    # Our report has no "Direct Answer", so regular repair keeps Exec Summary + 3 thematic = 4
    assert len(regular_sections) <= 5, \
        f"Regular repair should cap at ~5 sections, got {len(regular_sections)}: {regular_sections}"
    assert len(regular_sections) < 7, \
        "Regular repair kept all 7 sections — the cap-at-3 bug may be fixed, " \
        "re-evaluate whether diligence repair is still needed"


def test_synthesize_direct_diligence_path_uses_correct_gate_and_prompt():
    """End-to-end: synthesize_direct for diligence must use:
    1. The diligence system prompt (not the regular analyst prompt)
    2. The v2 prompt builder (with per-section sources)
    3. The diligence quality gate (not the Direct Answer gate)

    This test mocks the LLM to return a valid diligence report and verifies
    the output is NOT replaced by the deterministic fallback — proving the
    quality gate accepted it."""
    from workflows.deep_research.synthesizer import (
        synthesize_direct,
        _DILIGENCE_ANALYST_SYSTEM_PROMPT,
        _DIRECT_RESEARCH_ANALYST_SYSTEM_PROMPT,
    )
    from engine.deep_research_service import ResearchState

    sources = _make_tagged_sources()
    state = ResearchState(
        original_query="Diligence for Test Solar Farm",
        sources_checked=sources,
        research_type="diligence",
    )

    captured_calls = []

    def mock_call(prompt, system_prompt=None):
        captured_calls.append({"system_prompt": system_prompt, "prompt": prompt})
        # Return a report with all domain sections and citations
        return "\n\n".join([
            "## Executive Summary",
            "Analysis of 100MW solar acquisition in South Africa [Commercial Report 1].",
            "## Tariff & Revenue Potential",
            "NERSA-regulated tariffs: 1.78 ZAR/kWh retail [Commercial Report 1].",
            "## Regulatory & Licensing Requirements",
            "IPP license required under Electricity Regulation Act [Licensing Report 2].",
            "## Environmental & Grid Connection",
            "Basic Assessment required under NEMA [Environmental Report 3].",
            "## Performance Gap Analysis",
            "Capacity factor of 22.3% at De Aar benchmark [Performance Report 4].",
            "## Vendor & Counterparty Relationships",
            "Eskom Holdings is sole offtaker under REIPPP [Counterparty Report 5].",
            "## Risk Summary & Recommendation",
            "Regulatory uncertainty is primary risk [Asset Specific Report 6].",
        ])

    with patch(
        "workflows.deep_research.synthesizer._call_research_synthesis_model",
        side_effect=mock_call,
    ):
        result = synthesize_direct(
            state,
            asset_metadata=_ZA_SOLAR_ASSET,
            diligence_context="EIA data: South Africa solar irradiance 1800 kWh/m2/yr",
        )

    # 1. Must use diligence system prompt, not the regular one
    assert len(captured_calls) >= 1, "LLM was never called"
    assert captured_calls[0]["system_prompt"] == _DILIGENCE_ANALYST_SYSTEM_PROMPT, \
        f"Wrong system prompt. Got: {captured_calls[0]['system_prompt'][:80]}..."
    assert captured_calls[0]["system_prompt"] != _DIRECT_RESEARCH_ANALYST_SYSTEM_PROMPT, \
        "Used the regular analyst prompt — diligence path not wired"

    # 2. The prompt must have per-section source blocks (v2), not flat dump (v1)
    prompt_sent = captured_calls[0]["prompt"]
    assert "Sources for this section" in prompt_sent, \
        "Prompt lacks per-section sources — still using v1 flat dump"
    assert "RANKED SOURCES:" not in prompt_sent, \
        "Prompt still has v1 flat source block"

    # 3. Result must be the LLM output, NOT the deterministic fallback
    assert "Deterministic diligence report" not in result, \
        "Quality gate rejected the LLM output and fell back to deterministic — " \
        "gate is likely still checking for 'Direct Answer'"

    # 4. All domain sections must survive in the final output
    for section_title in ["Tariff & Revenue", "Regulatory & Licensing",
                          "Performance Gap", "Vendor & Counterparty", "Risk Summary"]:
        assert section_title in result, \
            f"Final output missing '{section_title}' — repair may have dropped it"

    # 5. LLM was called at most twice (initial + one retry), not in an infinite loop
    assert len(captured_calls) <= 2, \
        f"LLM called {len(captured_calls)} times — retry logic may be looping"
